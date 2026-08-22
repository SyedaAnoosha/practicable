"""Render a spreadsheet's real first screenful as a preview image (W4-R1, ledger row 24).

Why this is not a mockup
------------------------
Every cell drawn here is read out of the actual `.xlsx` in Storage. Nothing is invented,
padded or rearranged: the column order, the header row and the cell text are the file's
own. What is *not* reproduced is the source workbook's visual styling — openpyxl cannot
resolve theme-indexed fills to RGB without the theme part, and a half-reproduced style
looks like a rendering bug rather than the artefact. So the content is drawn in this
platform's own table treatment (theme.css §12.2 light tokens), which is also how the same
data would look inside a lesson on this site.

That is a deliberate, stated trade: **the data is the artefact's, the styling is ours.**
The alt text says so, so a buyer is never told they are looking at the file's own design.

Used by `generate_template_previews.py`. Kept separate because the PDF path rasterises a
page and this path composes one — two genuinely different jobs behind one interface.
"""
from __future__ import annotations

from io import BytesIO

# theme.css §12.2 — the light plane, the only plane a preview JPEG can be in (an image
# cannot respond to `prefers-color-scheme`, and light is the neutral ground the gallery
# thumbnails sit on in both themes).
INK = (28, 23, 18)             # --foreground  #1C1712
MUTED_INK = (109, 96, 82)      # --muted-foreground
CARD = (255, 255, 255)         # --card
BAND = (250, 247, 240)         # --card-2, the zebra band
HEADER_BG = (243, 238, 228)    # a header rail one step past --card-2
BORDER = (226, 219, 206)       # --border
GOLD = (154, 122, 60)          # --gold-strong, the header rule only

# Enough of the sheet to judge the artefact, not enough to be the artefact.
MAX_ROWS = 16

# Enough empty rows under a blank form's headers to read as a table rather than a single
# stripe, without pretending there is content in them.
BLANK_FORM_ROWS = 6
MAX_COLS = 6

CELL_PAD_X = 10
ROW_HEIGHT = 30
HEADER_HEIGHT = 34
MARGIN = 24
TITLE_HEIGHT = 46
FOOTER_HEIGHT = 30

# Column widths are the workbook's own, converted from Excel character units to pixels
# (~7px per character at the default font), then clamped so one prose column cannot
# squeeze the rest off the canvas.
PX_PER_CHAR = 7
MIN_COL_PX = 90
MAX_COL_PX = 300
DEFAULT_COL_PX = 150


def _load_font(size: int, bold: bool = False):
    """A real UI font where one is installed, Pillow's bitmap default otherwise. The
    fallback is deliberately allowed to look plain rather than crashing the run: a
    preview that renders is worth more than one that doesn't."""
    from PIL import ImageFont

    candidates = (
        ["segoeuib.ttf", "arialbd.ttf", "calibrib.ttf", "DejaVuSans-Bold.ttf"]
        if bold
        else ["segoeui.ttf", "arial.ttf", "calibri.ttf", "DejaVuSans.ttf"]
    )
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # A computed average arrives as 3.3333333333333335. Rendering the full float in
        # a preview reads as a defect in the artefact rather than as binary floating
        # point, so it is shown the way a spreadsheet would display it.
        return str(int(value)) if value.is_integer() else ("%.2f" % value).rstrip("0").rstrip(".")
    return str(value).replace("\n", " ").strip()


def _truncate(draw, text: str, font, max_px: int) -> str:
    """Ellipsis on overflow — a clipped word mid-glyph reads as a rendering fault."""
    if not text:
        return ""
    if draw.textlength(text, font=font) <= max_px:
        return text
    ellipsis = "…"
    while text and draw.textlength(text + ellipsis, font=font) > max_px:
        text = text[:-1]
    return (text.rstrip() + ellipsis) if text else ""


def _find_header_row(rows: list[list], max_scan: int = 6) -> int:
    """The first row where most cells are short, non-empty labels. Real workbooks open
    with a merged title banner and blank spacer rows above the actual header, and
    starting the table at row 1 would render those as a column of empty boxes."""
    best, best_score = 0, -1
    for i, row in enumerate(rows[:max_scan]):
        filled = [c for c in row if _cell_text(c)]
        if len(filled) < 2:
            continue
        shortish = sum(1 for c in filled if len(_cell_text(c)) <= 40)
        score = len(filled) + shortish
        if score > best_score:
            best, best_score = i, score
    return best


# Sheets that are legal/boilerplate rather than the artefact. Matched on the sheet NAME,
# because that is how these workbooks label them ("- Disclaimer -"). A picture of a
# third-party disclaimer is not a preview of what the buyer gets.
BOILERPLATE_SHEET_MARKERS = ("disclaimer", "terms", "license", "licence", "copyright", "about")


def _is_boilerplate_sheet(name: str) -> bool:
    lowered = name.lower().strip(" -_")
    return any(m in lowered for m in BOILERPLATE_SHEET_MARKERS)


def render_sheet_previews(blob: bytes, limit: int = 2) -> list[tuple[bytes, str, int]]:
    """Up to `limit` previews, one per content sheet.

    W4-R1 asks for at least two previews. This returns a second one only where a second
    *content* sheet actually exists — see this module's own note. Fewer than `limit` is a
    correct answer for a single-sheet workbook, not a failure to be padded.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(blob), read_only=True)
        names = [
            ws.title for ws in wb.worksheets
            if getattr(ws, "sheet_state", "visible") == "visible"
            and not _is_boilerplate_sheet(ws.title)
        ]
        wb.close()
    except Exception:
        names = []

    out: list[tuple[bytes, str, int]] = []
    for name in names[:limit]:
        rendered = render_sheet_preview(blob, sheet_name=name)
        if rendered is not None:
            out.append(rendered)
    return out


def render_sheet_preview(blob: bytes, sheet_name: str | None = None) -> tuple[bytes, str, int] | None:
    """First worksheet's opening screenful as JPEG bytes.

    Returns `(jpeg, sheet_name, columns_shown)` or None if the workbook cannot be read
    or holds nothing worth showing — in which case the caller writes no preview at all.
    """
    from openpyxl import load_workbook
    from PIL import Image, ImageDraw

    try:
        wb = load_workbook(BytesIO(blob), read_only=False, data_only=True)
    except Exception:
        return None

    try:
        if sheet_name is not None:
            sheet = next((ws for ws in wb.worksheets if ws.title == sheet_name), None)
        else:
            sheet = next(
                (ws for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"),
                None,
            )
        if sheet is None:
            return None

        raw = [list(r) for r in sheet.iter_rows(max_row=40, max_col=20, values_only=True)]
        if not raw:
            return None

        # Drop wholly empty leading rows and columns — the source files start at B2, and
        # rendering column A as an empty stripe would look like a bug.
        while raw and not any(_cell_text(c) for c in raw[0]):
            raw.pop(0)
        if not raw:
            return None

        col_used = [
            any(_cell_text(row[i]) for row in raw if i < len(row))
            for i in range(max(len(r) for r in raw))
        ]
        keep = [i for i, used in enumerate(col_used) if used][:MAX_COLS]
        if not keep:
            return None

        header_idx = _find_header_row(raw)
        headers = [_cell_text(raw[header_idx][i]) if i < len(raw[header_idx]) else "" for i in keep]
        body_rows = [
            [_cell_text(r[i]) if i < len(r) else "" for i in keep]
            for r in raw[header_idx + 1 :]
        ]
        body_rows = [r for r in body_rows if any(r)][:MAX_ROWS]

        # A blank form (headers, no data) is not "nothing to show": its column structure
        # IS the artefact, and "what fields does this track?" is the question a buyer of a
        # fill-in-yourself template actually has. Render the headers over a few empty rows
        # and let the footer say plainly that it is a blank form.
        is_blank_form = not body_rows
        if is_blank_form:
            if not any(headers):
                # Neither headers nor data — this genuinely has nothing to show.
                return None
            body_rows = [["" for _ in keep] for _ in range(BLANK_FORM_ROWS)]

        # A blank checklist has real headers over entirely empty cells. Keeping those
        # columns is faithful to the file but renders as a wall of whitespace that
        # undersells the artefact, so a column with a header and no data in any shown
        # row is dropped — unless dropping would leave fewer than two columns, in which
        # case the emptiness IS the artefact (a fill-in-yourself form) and is kept.
        populated = [
            i for i in range(len(keep))
            if any(row[i] for row in body_rows)
        ]
        dropped_empty = 0
        if not is_blank_form and len(populated) >= 2 and len(populated) < len(keep):
            dropped_empty = len(keep) - len(populated)
            headers = [headers[i] for i in populated]
            body_rows = [[row[i] for i in populated] for row in body_rows]
            keep = [keep[i] for i in populated]

        # Column widths from the workbook, clamped.
        widths: list[int] = []
        for i in keep:
            letter = sheet.cell(row=1, column=i + 1).column_letter
            dim = sheet.column_dimensions.get(letter)
            px = int((dim.width or 0) * PX_PER_CHAR) if dim and dim.width else DEFAULT_COL_PX
            widths.append(max(MIN_COL_PX, min(MAX_COL_PX, px)))

        table_w = sum(widths)
        width = table_w + MARGIN * 2
        height = MARGIN * 2 + TITLE_HEIGHT + HEADER_HEIGHT + ROW_HEIGHT * len(body_rows) + FOOTER_HEIGHT

        img = Image.new("RGB", (width, height), CARD)
        draw = ImageDraw.Draw(img)

        f_title = _load_font(19, bold=True)
        f_head = _load_font(13, bold=True)
        f_body = _load_font(13)
        f_foot = _load_font(12)

        # Title: the sheet's own name, which is the artefact's own label for this view.
        draw.text((MARGIN, MARGIN), sheet.title, font=f_title, fill=INK)
        rule_y = MARGIN + TITLE_HEIGHT - 12
        draw.rectangle([MARGIN, rule_y, MARGIN + 56, rule_y + 2], fill=GOLD)

        y = MARGIN + TITLE_HEIGHT
        draw.rectangle([MARGIN, y, MARGIN + table_w, y + HEADER_HEIGHT], fill=HEADER_BG)
        x = MARGIN
        for w, text in zip(widths, headers):
            draw.text(
                (x + CELL_PAD_X, y + (HEADER_HEIGHT - 15) // 2),
                _truncate(draw, text, f_head, w - CELL_PAD_X * 2),
                font=f_head,
                fill=INK,
            )
            x += w
        y += HEADER_HEIGHT
        draw.line([MARGIN, y, MARGIN + table_w, y], fill=BORDER, width=1)

        for r_i, row in enumerate(body_rows):
            if r_i % 2 == 1:
                draw.rectangle([MARGIN, y, MARGIN + table_w, y + ROW_HEIGHT], fill=BAND)
            x = MARGIN
            for w, text in zip(widths, row):
                draw.text(
                    (x + CELL_PAD_X, y + (ROW_HEIGHT - 15) // 2),
                    _truncate(draw, text, f_body, w - CELL_PAD_X * 2),
                    font=f_body,
                    fill=INK if text else MUTED_INK,
                )
                x += w
            y += ROW_HEIGHT
            draw.line([MARGIN, y, MARGIN + table_w, y], fill=BORDER, width=1)

        # Vertical rules, drawn last so they sit above the zebra bands.
        x = MARGIN
        top = MARGIN + TITLE_HEIGHT
        for w in widths[:-1]:
            x += w
            draw.line([x, top, x, y], fill=BORDER, width=1)
        draw.rectangle([MARGIN, top, MARGIN + table_w, y], outline=BORDER, width=1)

        # The footer is load-bearing, not decoration: it stops the image from implying
        # the sheet ends where the preview does.
        total_rows = max(0, sheet.max_row - header_idx - 1)
        if is_blank_form:
            # Never imply data the file does not contain.
            note = "Blank form - " + str(len(headers)) + " column"
            note += "" if len(headers) == 1 else "s"
            note += " to fill in"
        else:
            note = "Preview of the first " + str(len(body_rows)) + " rows"
            if total_rows > len(body_rows):
                note += " of " + str(total_rows)
        note += " · " + str(len(wb.worksheets)) + (" sheet" if len(wb.worksheets) == 1 else " sheets") + " in the file"
        if dropped_empty:
            # Never let the preview imply the file has fewer columns than it does.
            note += " · " + str(dropped_empty) + (
                " blank column hidden" if dropped_empty == 1 else " blank columns hidden"
            )
        draw.text((MARGIN, y + 9), note, font=f_foot, fill=MUTED_INK)

        buf = BytesIO()
        img.save(buf, format="JPEG", quality=88, optimize=True, progressive=True)
        return buf.getvalue(), sheet.title, len(keep)
    finally:
        wb.close()
